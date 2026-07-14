# =============================================================================
# tabs/train_tab.py
#
# PURPOSE
# -------
# Renders the "Train & Validate" tab of the SINDy Expert System.
# Responsibilities:
#   1. Let the user pick a dataset (pre-set system or custom CSV upload).
#   2. Run an automatic data-analysis heuristic ("AI Suggester") that
#      recommends starting values for polynomial degree and sparsity
#      threshold based on linearity / periodicity / noise level.
#   3. Fit a SINDy model on a random train/validation split.
#   4. Show the fitted trajectory, a leaderboard of all past runs, and
#      residual diagnostics (time-domain, frequency-domain, and
#      true-vs-predicted scatter) so the user can judge model quality.
#   5. Allow viewing/deleting any past run from the leaderboard.
# =============================================================================

from bokeh.models import (ColumnDataSource, Slider, Div, Button,
                          Select, DataTable, TableColumn, HTMLTemplateFormatter, FileInput, TextInput)
from bokeh.layouts import column, row
from bokeh.plotting import figure
import pandas as pd
import numpy as np
import os
import copy
import base64
import io
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# AUTO-EXPORT: on-disk leaderboard log
#
# Mirrors the in-app "TRAINING HISTORY" DataTable (Section 4) 1:1, column
# for column, but persisted to an .xlsx file on every successful Train click
# so results survive app restarts / server reloads. Kept as a small,
# self-contained block (no dependency on Bokeh widgets) so it can be reused
# or moved to its own module later without touching the tab-building code.
# =============================================================================

# ── Backend switch: "excel" (default, on-disk .xlsx) or "gsheet" (Google
#    Sheets, via a service account). Set with an env var so no code change
#    is needed to switch: SINDY_LOG_BACKEND=gsheet
LOG_BACKEND = os.environ.get("SINDY_LOG_BACKEND", "excel").strip().lower()

LOG_DIR = "logs"
LOG_FILE_PATH = os.path.join(LOG_DIR, "training_history_log.xlsx")
LOG_SHEET_NAME = "Training History"


# Same column set/order as the on-screen leaderboard (`columns` in Section 4).
# "Data File" records which dataset/system produced the run, for context.
LOG_HEADERS = [
    "Run #", "Data File", "Split Type", "Library", "Degree", "Threshold",
    "Train R2 (dX)", "Train RMSE (dX)", "Train MAE (dX)",
    "Val R2 (dX)", "Val RMSE (dX)", "Val MAE (dX)",
    "RMSE Diff", "Identified Equations",
]
# Column widths chosen to keep roughly the same visual proportions as the
# in-app DataTable's pixel widths, scaled down to Excel's character units.
LOG_COL_WIDTHS = [8, 22, 16, 12, 9, 11, 14, 15, 14, 14, 15, 14, 12, 65]


_HEADER_FILL = PatternFill(start_color="1F4E79",
                           end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_ZEBRA_FILL = PatternFill(start_color="F2F2F2",
                          end_color="F2F2F2", fill_type="solid")
_THIN_BORDER = Border(*(Side(style="thin", color="BFBFBF"),) * 4)


def _strip_html(html_text):
    """
    Turn the HTML-formatted equation string used in the in-app DataTable
    (`formatted_eqs_html`, e.g. "<b>(1)</b> dx0/dt = ...<br>...") into
    plain, newline-separated text suitable for an Excel cell.
    """
    text = re.sub(r"<br\s*/?>", "\n", html_text)
    text = re.sub(r"<[^>]+>", "", text)
    return text.strip()


def _ensure_log_workbook():
    """
    Open the log workbook if it already exists, otherwise create it with a
    styled header row (same dark-blue header used across the app's other
    tables) and sensible column widths. Called once per export so the file
    is always self-healing even if it was deleted/moved between runs.
    """
    os.makedirs(LOG_DIR, exist_ok=True)

    if os.path.exists(LOG_FILE_PATH):
        wb = load_workbook(LOG_FILE_PATH)
        if LOG_SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(LOG_SHEET_NAME)
        else:
            return wb
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = LOG_SHEET_NAME

    ws.append(LOG_HEADERS)
    for col_idx in range(1, len(LOG_HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER

    for col_idx, width in enumerate(LOG_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    return wb


def _export_to_excel(row_values):
    wb = _ensure_log_workbook()
    ws = wb[LOG_SHEET_NAME]
    row_idx = ws.max_row + 1

    ws.append(row_values)

    # Zebra striping + wrap text, matching the in-app table's look.
    if row_idx % 2 == 0:
        for col_idx in range(1, len(LOG_HEADERS) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = _ZEBRA_FILL
    for col_idx in range(1, len(LOG_HEADERS) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = _THIN_BORDER
        if col_idx == len(LOG_HEADERS):
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(LOG_FILE_PATH)



def export_run_to_log(run_id, data_file, split_type, lib, poly_degree, threshold,
                      train_r2, train_rmse, train_mae,
                      val_r2, val_rmse, val_mae, rmse_diff,
                      equations_html):
    """
    Append one completed training run to the leaderboard log — either an
    on-disk .xlsx (default) or a Google Sheet, depending on LOG_BACKEND.
    Called once per successful Train click (see on_train_click, Section 7),
    right after the run is added to the in-app DataTable, so the two stay
    in lockstep. Never raises — a logging failure must not break training.
    """
    row_values = [
        run_id, data_file, split_type, lib, poly_degree, threshold,
        round(train_r2, 4), round(train_rmse, 6), round(train_mae, 6),
        round(val_r2, 4), round(val_rmse, 6), round(val_mae, 6),
        round(rmse_diff, 6), _strip_html(equations_html),
    ]
    try:
        _export_to_excel(row_values)
    except Exception as e:
        # Logging must never take down the training flow.
        target = "Google Sheet" if LOG_BACKEND == "gsheet" else LOG_FILE_PATH
        print(f"[WARN] Could not write training log to {target}: {e}")


def reset_log():
    """
    Wipe the training history log back to just the header row. Works for
    whichever backend is currently configured (LOG_BACKEND).
    """
    if os.path.exists(LOG_FILE_PATH):
        os.remove(LOG_FILE_PATH)
        print(f"[OK] Deleted {LOG_FILE_PATH}")
        _ensure_log_workbook().save(LOG_FILE_PATH)
        print(f"[OK] Recreated empty log at {LOG_FILE_PATH}")


if __name__ == "__main__":
    # ── Terminal command to reset the log ───────────────────────────────
    #   python tabs/train_tab.py --reset-log
    # Works against whichever backend SINDY_LOG_BACKEND points to.
    import argparse
    parser = argparse.ArgumentParser(
        description="SINDy training log utilities")
    parser.add_argument("--reset-log", action="store_true",
                        help="Clear the training history log and recreate it with just the header row.")
    args = parser.parse_args()
    if args.reset_log:
        reset_log()
    else:
        parser.print_help()


def train_tab_layout(engine, trained_model_storage):
    """
    Build the Bokeh layout for the Train & Validate tab.

    Parameters
    ----------
    engine : SINDyEngine
        Shared engine instance (holds the pySINDy model, fit/simulate/
        diagnostics methods). Same instance is passed to Test/Predict tabs
        so that trained models can be reused across tabs.
    trained_model_storage : dict
        Shared in-memory store: {run_id: {model_instance, metrics, plot_data,
        diagnostics, ...}}. Acts as the "database" for the leaderboard and
        is read by the Test/Predict tabs to let the user pick a trained run.

    Returns
    -------
    bokeh.layouts.column
        The complete tab layout, ready to be added to a Bokeh document.
    """

    # =========================================================================
    # SECTION 1 — AI SUGGESTER
    # Heuristic data analysis that recommends a starting library / degree /
    # threshold before the user has to guess hyperparameters manually.
    # =========================================================================

    def analyze_data_linearity(df):
        """
        Analyze an uploaded/selected dataset and suggest SINDy hyperparameters.

        Strategy
        --------
        1. Estimate dX/dt via Savitzky-Golay smoothing + finite differences
           (smoothing reduces the amplification of noise inherent to
           numerical differentiation).
        2. Fit polynomial regressions of degree 1, 2, and 3 from X -> dX/dt
           and compare their R² scores. The *smallest* degree that gives a
           meaningfully better fit than the previous degree is selected —
           this avoids over-suggesting complexity when a simpler model
           already explains the dynamics well (Occam's razor).
        3. Run an FFT-based periodicity check (dominant peak vs mean
           spectral energy) — currently informational only; see the note
           in Step 6 for why the library suggestion itself is not directly
           applied to the UI.
        4. Estimate the noise floor from the high-frequency tail of the FFT
           and map it to a suggested sparsity threshold.

        Returns
        -------
        tuple(str, int, float, str)
            (suggested_library, suggested_degree, suggested_threshold, reason_text)
            reason_text is a human-readable explanation shown in the UI.
        """
        try:
            t = df.iloc[:, 0].values
            X = df.iloc[:, 1:].values
            n_vars = X.shape[1]
            dt = np.mean(np.diff(t)) if len(t) > 1 else 0.1

            # ── 1. Calculate derivative ──────────────────────────────
            # Savitzky-Golay smoothing before differentiating: this is
            # critical because raw finite differences amplify sensor/CSV
            # noise, which would otherwise bias the linearity test below.
            from scipy.signal import savgol_filter
            dXdt = np.zeros_like(X)
            window = min(11, len(t) // 10 * 2 + 1)
            window = max(window, 5)
            for i in range(n_vars):
                smoothed = savgol_filter(
                    X[:, i], window_length=window, polyorder=3)
                dXdt[:, i] = np.gradient(smoothed, dt)

            # ── 2. Compare R² of degree 1, 2, 3 ────────────────────────
            # Fit an ordinary polynomial regression (not SINDy/STLSQ) purely
            # as a fast proxy to gauge how nonlinear the system "looks".
            from sklearn.preprocessing import PolynomialFeatures
            from sklearn.linear_model import LinearRegression
            from sklearn.metrics import r2_score

            r2_scores = {}
            for deg in [1, 2, 3]:
                poly = PolynomialFeatures(degree=deg, include_bias=True)
                X_poly = poly.fit_transform(X)
                lr = LinearRegression(fit_intercept=False).fit(X_poly, dXdt)
                r2_scores[deg] = r2_score(dXdt, lr.predict(X_poly),
                                          multioutput='uniform_average')

            r2_linear = r2_scores[1]
            r2_deg2 = r2_scores[2]
            r2_deg3 = r2_scores[3]

            # ── 3. Choose minimal degree that satisfies R² ──────────────────
            # If degree=1 is already sufficient → treat as a linear system.
            if r2_linear >= 0.85:
                sug_degree = 1
            # If degree=2 meaningfully improves over degree=1 → nonlinear (quadratic-ish).
            elif r2_deg2 - r2_linear >= 0.05:
                sug_degree = 2
            # If degree=3 meaningfully improves over degree=2 → higher-order nonlinearity.
            elif r2_deg3 - r2_deg2 >= 0.05:
                sug_degree = 3
            # No degree gives a meaningful improvement → default back to linear
            # (a weakly-nonlinear system may simply be indistinguishable from
            # noise at this sampling rate/noise level).
            else:
                sug_degree = 1

            # ── 4. FFT — detect periodicity ──────────────────────────────
            # A single dominant peak that is >5x the mean spectral amplitude
            # indicates a strongly oscillatory/periodic signal.
            is_periodic = False
            for i in range(n_vars):
                # remove DC offset before FFT
                signal = X[:, i] - np.mean(X[:, i])
                fft_vals = np.abs(np.fft.rfft(signal))
                peaks = fft_vals[1:]  # skip the DC bin
                if len(peaks) > 0:
                    if np.max(peaks) > 5 * np.mean(peaks):
                        is_periodic = True
                        break

            # ── 5. Noise estimate → suggest threshold ────────────────────
            # Approximate the noise floor as the median amplitude of the
            # top 20% highest frequencies (where genuine physical signal
            # content is usually negligible for smooth trajectories).
            noise_estimates = []
            for i in range(n_vars):
                amp = np.abs(np.fft.rfft(X[:, i])) / len(t)
                high = np.sort(amp)[-max(1, int(len(amp) * 0.2)):]
                noise_estimates.append(float(np.median(high)))
            noise_level = float(np.mean(noise_estimates))

            if noise_level < 0.01:
                sug_threshold = 0.05
            elif noise_level < 0.05:
                sug_threshold = 0.10
            else:
                sug_threshold = 0.20

            # ── 6. Library suggestion (informational only — NOT applied to UI) ──
            # NOTE: Earlier testing (see project history) showed that Fourier/
            # Combined suggestions frequently misfire on purely polynomial
            # systems that merely *look* oscillatory (e.g. coupled spring-mass),
            # because the peak-ratio heuristic can't distinguish "oscillatory
            # data" from "equations that actually contain sin/cos terms".
            # We therefore compute sug_library for display/reasoning purposes
            # only; apply_suggestion() below deliberately does NOT set
            # library_select.value from this result. Only degree and
            # threshold are auto-applied.
            if is_periodic and r2_linear < 0.92:
                sug_library = "Combined"
            elif is_periodic and r2_linear >= 0.92:
                sug_library = "Fourier"
            else:
                sug_library = "Polynomial"

            reason = f"Degree: {sug_degree}; Threshold: {sug_threshold}; Noise ≈ {noise_level:.4f}."
            return sug_library, sug_degree, sug_threshold, reason

        except Exception as e:
            # Fail-safe defaults so a bad/edge-case CSV never blocks the user
            # from proceeding to manual configuration.
            return "Polynomial", 1, 0.10, f"Error analyzing data: {e}"

    def apply_suggestion(df, prefix_msg=""):
        """
        Run analyze_data_linearity() and push the result into the UI:
        updates poly_s / thr_s slider values and displays the reasoning
        text in upload_status. (library is intentionally NOT auto-applied,
        see note in Step 6 above.)
        """
        lib, deg, thr, reason = analyze_data_linearity(df)
        poly_s.value = deg
        thr_s.value = thr
        upload_status.text = f"{prefix_msg}<br><b>Suggestion:</b> {reason}"

    # =========================================================================
    # SECTION 2 — DATA SOURCE SELECTION
    # Dropdown for pre-set systems + custom CSV upload widget.
    # =========================================================================

    system_options = [
        ("cs_train_data.csv",   "Coupled Spring-Mass (Pre-set)"),
        ("vanderpol_train.csv", "Van der Pol Oscillator (Pre-set)"),
        ("custom_upload",       "Upload your own data")
    ]

    file_select = Select(title="SELECT SYSTEM", options=system_options,
                         value="cs_train_data.csv")
    # value -> friendly label lookup, used to populate the "Data File"
    # column in the leaderboard (e.g. "cs_train_data.csv" -> "Coupled
    # Spring-Mass (Pre-set)").
    _SYSTEM_LABELS = dict(system_options)

    # File upload widget — hidden until the user picks "Upload your own data".
    file_input = FileInput(accept=".csv", visible=False)
    upload_status = Div(
        text="", styles={'color': "#247008", 'font-size': '13px'})
    # caches the last base64 payload (currently informational)
    _upload_buffer = {'data': None}

    def on_file_select_change(attr, old, new):
        """
        Toggle the upload widget visibility and, for pre-set systems,
        immediately load the CSV and run the AI Suggester so the sliders
        are pre-filled before the user even presses Train.
        """
        if new == "custom_upload":
            file_input.visible = True
            upload_status.text = "Please upload a CSV with columns: t, x1, x2..."
        else:
            file_input.visible = False
            path = os.path.join('data', new)
            if os.path.exists(path):
                df = pd.read_csv(path).astype(np.float64)
                apply_suggestion(df, f"<b>Selected system file: {new}</b>")
            else:
                upload_status.text = f"⚠ Pre-set file not found at {path}"

    file_select.on_change('value', on_file_select_change)

    def upload_to_local_drive(attr, old, new):
        """
        Callback fired when FileInput receives a new file. Bokeh delivers
        the file content as a base64 string in `new`. We decode it into a
        DataFrame purely to run the AI Suggester immediately (the actual
        training callback re-decodes file_input.value independently — see
        on_train_click — so this decode here is "preview only").
        """
        if not new:
            return
        _upload_buffer['data'] = new  # cache base64 payload
        try:
            decoded = base64.b64decode(new)
            f = io.BytesIO(decoded)
            df = pd.read_csv(f).astype(np.float64)
            apply_suggestion(df, "Custom file uploaded successfully!")
        except Exception as e:
            upload_status.text = f"⚠ Error processing uploaded file: {e}"

    file_input.on_change('value', upload_to_local_drive)

    # =========================================================================
    # SECTION 3 — MODEL CONFIGURATION CONTROLS
    # Library type, train/validation split, polynomial degree, sparsity
    # threshold, and the Train button.
    # =========================================================================

    library_select = Select(title="LIBRARY",
                            options=["Polynomial", "Fourier", "Combined"],
                            value="Polynomial")

    # Single slider controls the split; validation % is always 100 - train%.
    # (This UX choice avoids the earlier bug where two independent sliders
    # could be set to sum to less/more than 100%.)
    train_s = Slider(start=10, end=90, value=60, step=5,
                     title="Train/Validation Split")
    def on_train_s_change(attr, old, new):
        """Keep the human-readable split label in sync with the slider."""
        train_s.title = f"SPLIT: TRAIN {new}% | VALIDATION {100 - new}%"
        
    train_s.on_change('value', on_train_s_change)
    # Initialize title immediately
    on_train_s_change(None, None, train_s.value)
    train_s.show_value = False
    split_select = Select(
        title="SPLIT TYPE",
        value="Random Sampling",
        options=[
            "Random Sampling",
            "Time-based",
            "Random Block",
        ],
        width=150,
    )



    poly_s = Slider(start=1, end=5,     value=1,
                    step=1,     title="DEGREE / HARMONICS")
    thr_s = Slider(start=0.0, end=0.5, value=0.1,
                   step=0.005, title="SPARSITY THRESHOLD")
    # ── Manual threshold input ──────────────────────────────────────────
    # Some systems turned out to be very sensitive to the exact threshold
    # value — the 0.005 slider step is too coarse for fine-tuning (e.g.
    # 0.0347 vs 0.035). This TextInput lets the user type an exact value;
    # it's two-way synced with thr_s so either control can drive the other.
    thr_input = TextInput(
        value=f"{thr_s.value:.4f}", title="Or type exact threshold:", width=150)

    # re-entrancy guard to prevent infinite update loops
    _thr_syncing = [False]

    def on_thr_slider_change(attr, old, new):
        """Slider moved → push the new value into the text box."""
        if _thr_syncing[0]:
            return
        _thr_syncing[0] = True
        thr_input.value = f"{new:.4f}"
        _thr_syncing[0] = False

    def on_thr_input_change(attr, old, new):
        """
        Text box edited → validate and push into the slider. Falls back
        silently to the last valid value if the typed text isn't a
        parseable, in-range number (e.g. mid-typing state like "0." or
        "-"), so the app never crashes on invalid manual input.
        """
        if _thr_syncing[0]:
            return
        try:
            val = float(new)
        except ValueError:
            return

        val = max(thr_s.start, min(thr_s.end, val))  # clamp to valid range

        _thr_syncing[0] = True
        thr_s.value = val
        thr_input.value = f"{val:.4f}"
        _thr_syncing[0] = False

    thr_s.on_change('value', on_thr_slider_change)
    thr_input.on_change('value', on_thr_input_change)
    btn_train = Button(label="TRAIN", button_type="primary",
                       height=50, width=100)

    # =========================================================================
    # SECTION 4 — HISTORY TABLE (LEADERBOARD)
    # Stores one row per training run with all validation metrics and the
    # discovered equations. Selecting a row re-renders that run's plots.
    # =========================================================================

    # Custom HTML template so multi-line equation strings wrap nicely
    # inside the DataTable cell instead of being clipped.
    eqn_template = """
    <div style="white-space: normal; word-wrap: break-word; line-height: 1.5;
                padding: 8px 0; font-family: 'Courier New', monospace;
                font-size: 12px; color: #00000;">
        <%= value %>
    </div>
    """
    eqn_formatter = HTMLTemplateFormatter(template=eqn_template)

    # Small HTML template for the merged Train/Val metric cells (Section 4
    # "compact" view) — packs R²/RMSE/MAE into 3 short lines instead of 3
    # separate wide columns, so the leaderboard needs a lot less horizontal
    # space per run.
    metrics_template = """
    <div style="white-space: normal; line-height: 1.4; padding: 4px 0;
                font-family: 'Courier New', monospace; font-size: 11px;">
        <%= value %>
    </div>
    """
    metrics_formatter = HTMLTemplateFormatter(template=metrics_template)

    def _fmt_metrics_html(label, color, r2, rmse, mae):
        return (
            f"<b style='color:{color};'>{label}</b><br>"
            f"R²: {r2:.4f}<br>RMSE: {rmse:.6f}<br>MAE: {mae:.6f}"
        )

    # "system" records which dataset/file produced the run (pre-set name or
    # the uploaded filename) so old runs stay interpretable at a glance.
    source_history = ColumnDataSource(data=dict(
        run=[], system=[], split=[], lib=[], poly=[], thr=[],
        train_metrics=[], val_metrics=[],
        rmse_diff=[], equations=[]
    ))

    # Train/Val each collapse into a single merged HTML cell (R²+RMSE+MAE
    # stacked) instead of 6 separate wide numeric columns.
    columns = [
        TableColumn(field="run",    title="Run #",      width=100),
        TableColumn(field="system", title="Data File",  width=300),
        TableColumn(field="split",  title="Split Type", width=300),
        TableColumn(field="lib",    title="Library",    width=200),
        TableColumn(field="poly",   title="Degree",     width=200),
        TableColumn(field="thr",    title="Threshold",  width=200),
        TableColumn(field="train_metrics", title="Train Metrics",
                    width=200, formatter=metrics_formatter),
        TableColumn(field="val_metrics",   title="Val Metrics",
                    width=200, formatter=metrics_formatter),
        TableColumn(field="rmse_diff", title="RMSE Diff", width=200),
        TableColumn(field="equations", title="Identified Equations",
                    width=1000, formatter=eqn_formatter),
    ]

    history_table = DataTable(
        source=source_history, columns=columns,
        width=1400, height=400, row_height=200,
        index_position=None, background="#ffffff",
        sortable=True, selectable=True,
    )

    btn_delete = Button(label="DELETE",
                        button_type="danger", width=100, height=50)
    btn_delete.disabled = True  # default disable when there is no run

    def on_row_select(attr, old, new):
        """
        Fired when the user clicks a row in the leaderboard. Re-renders
        the main result plot AND the diagnostic plots using the stored
        data for that run — this is what lets users "time travel" back
        to any previous run without re-training.
        """
        # when choose run -> appear button, allows user to delete runs
        btn_delete.disabled = not bool(new)
        if not new:
            return
        run_id = source_history.data['run'][new[0]]
        if run_id in trained_model_storage:
            render_plot(run_id)
            diag = trained_model_storage[run_id].get('diagnostics')
            if diag:
                _render_diag_plots(diag)

    source_history.selected.on_change('indices', on_row_select)

    # =========================================================================
    # SECTION 5 — MAIN RESULT PLOT
    # Shows train/validation points scattered against the SINDy-simulated
    # trajectory for the currently-viewed run.
    # =========================================================================

    p = figure(title="Model Result", height=430, sizing_mode="stretch_width")
    # Empty invisible glyph forces Bokeh to allocate a renderer/legend slot
    # immediately, avoiding a "plot has zero renderers" warning on first load.
    p.scatter([], [], alpha=0)
    p.legend.click_policy = "hide"

    res_div = Div(
        text="<h3>Run Equations:</h3>",
        styles={'background': '#f8f9fa',
                'padding': '10px', 'border-radius': '5px'}
    )

    # =========================================================================
    # SECTION 6 — RESIDUAL DIAGNOSTIC PLOTS
    # Three complementary views of model quality on the derivative (dX/dt)
    # space, used to spot missing library terms or structured (non-random)
    # error that a single R²/RMSE number would hide.
    # =========================================================================

    # Plot 1: Residual vs Time
    # A well-fit model's residuals should look like structureless noise.
    # Visible trends/oscillations indicate the model is missing a term.
    p_resid = figure(
        title="Residual vs Time",
        sizing_mode="stretch_width",
        height=280,
        x_axis_label="Time",
        y_axis_label="Residual",
        toolbar_location=None,
    )
    p_resid.scatter([], [], alpha=0)

    # Plot 2: FFT of Residual
    # A dominant frequency peak in the residual spectrum means there is
    # still periodic structure in the error → the candidate library is
    # missing a sin/cos (or higher harmonic) term.
    p_fft = figure(
        title="Residual FFT (Frequency Content)",
        sizing_mode="stretch_width",
        height=280,
        x_axis_label="Frequency (Hz)",
        y_axis_label="Amplitude",
        toolbar_location=None,
    )
    p_fft.scatter([], [], alpha=0)

    # Plot 3: dX_true vs dX_predicted scatter
    # A perfect model places every point exactly on the y=x diagonal.
    # Systematic curvature/fanning indicates bias or heteroscedastic error.
    p_scatter = figure(
        title="dX True vs dX Predicted",
        sizing_mode="stretch_width",
        height=280,
        x_axis_label="dX Predicted",
        y_axis_label="dX True",
        toolbar_location=None,
    )
    p_scatter.scatter([], [], alpha=0)

    # Text summary shown above the 3 diagnostic plots (R², SNR, autocorrelation
    # per state variable).
    diag_stats_div = Div(
        text="<a>Run a training session to see diagnostics.</a>",
        styles={'font-size': '13px', 'color': '#00000'}
    )

    counter = [0]   # run counter — monotonically increasing, never reset
    # even after deletions (see project history: run IDs
    # are intentionally permanent to avoid ambiguity).
    view_div = Div(
        text="",
        styles={'color': '#7f8c8d', 'font-size': '13px', 'padding': '4px 0'}
    )

    def render_plot(run_id):
        """
        Redraw the main result plot (Section 5) from the stored plot_data
        of a given run. This is what allows switching between runs in the
        leaderboard without re-running the (potentially expensive) SINDy fit.
        """
        data = trained_model_storage[run_id]['plot_data']
        t, X = data['t'], data['X']
        train_idx = data['train_idx']
        val_idx = data['val_idx']
        x_sim_full = data['x_sim']

        p.renderers = []
        if p.legend and len(p.legend) > 0:
            p.legend.items = []

        # Train points in BLUE
        p.scatter(t[train_idx], X[train_idx, 0],
                  color="#1f77b4", alpha=0.4, size=4, legend_label="Train points")
        # Validation points in ORANGE
        p.scatter(t[val_idx], X[val_idx, 0],
                  color="#ff7f0e", alpha=0.4, size=4, legend_label="Val points")
        # SINDy-simulated trajectory in GREEN
        if x_sim_full is not None:
            p.line(t, x_sim_full[:, 0],
                   color="#2ecc71", line_width=2.5, legend_label="SINDy found")

        p.legend.click_policy = "hide"
        p.legend.location = "top_right"
        p.title.text = f"Model Result — Run #{run_id}"
        view_div.text = f"<b style='color:#2c3e50;'>👁 Viewing Run #{run_id}</b>"

    # Shared color palette for multi-variable diagnostic plots (cycles if
    # a system has more than 5 state variables).
    _DIAG_COLORS = ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd"]

    def _render_diag_plots(diag):
        """
        Populate the 3 diagnostic plots (Section 6) from a diagnostics dict
        produced by engine.compute_diagnostics().

        Expected `diag` structure:
            diag['t']            -> time array
            diag['residuals']    -> {var_name: residual array}
            diag['fft_freqs']    -> frequency axis (shared across variables)
            diag['fft_amps']     -> {var_name: FFT amplitude array}
            diag['dX_true']      -> {var_name: true derivative array}
            diag['dX_pred']      -> {var_name: predicted derivative array}
            diag['stats']        -> {var_name: {r2_dx, snr_db, autocorr}}

        FFT x-axis auto-scaling
        ------------------------
        We auto-scale the residual FFT x-axis to the frequency range that
        actually contains meaningful energy. This avoids two problems:
          1. Hardcoded ranges that only work for one specific system.
          2. Showing the full Nyquist range where most content is noise
             floor, making real peaks hard to see.
        """
        if not diag:
            return

        # Clear all 3 plots before redrawing.
        p_resid.renderers = []
        p_fft.renderers = []
        p_scatter.renderers = []
        if p_resid.legend:
            p_resid.legend.items = []
        if p_fft.legend:
            p_fft.legend.items = []
        if p_scatter.legend:
            p_scatter.legend.items = []

        var_names = list(diag['residuals'].keys())
        freqs = diag['fft_freqs']

        for idx, name in enumerate(var_names):
            color = _DIAG_COLORS[idx % len(_DIAG_COLORS)]

            # Plot 1 — Residual vs Time
            p_resid.line(
                diag['t'], diag['residuals'][name],
                color=color, line_width=1.5, alpha=0.8,
                legend_label=name,
            )

            # Plot 2 — FFT amplitude spectrum of the residual
            p_fft.line(
                freqs, diag['fft_amps'][name],
                color=color, line_width=1.5, alpha=0.8,
                legend_label=name,
            )

            # Plot 3 — dX_true vs dX_pred scatter
            p_scatter.scatter(
                diag['dX_pred'][name], diag['dX_true'][name],
                color=color, alpha=0.3, size=4,
                legend_label=name,
            )

        # ── Auto-scale FFT x-axis ──────────────────────────────────────────
        # Combine amplitude across all variables to find the global energy
        # envelope, then show only the range where at least one variable
        # has meaningful energy (> 1% of the global peak amplitude). This
        # generalizes to any system — slow biological oscillators, fast
        # mechanical systems, chaotic attractors — without any hardcoded
        # frequency limit.
        all_amps = np.concatenate([diag['fft_amps'][n] for n in var_names])
        max_amp = float(all_amps.max())

        if max_amp > 0:
            significant_indices = np.where(all_amps > 0.01 * max_amp)[0]

            if len(significant_indices) > 0:
                # all_amps is a concatenation of n_vars arrays each of
                # length n_freqs — map the flat index back onto the shared
                # frequency axis with a modulo.
                n_freqs = len(freqs)
                last_idx = int(significant_indices[-1]) % n_freqs
                f_max = float(freqs[last_idx])

                # 20% margin so the last visible peak isn't clipped at the edge.
                p_fft.x_range.end = f_max * 1.2
                p_fft.x_range.start = 0.0

        # Add a y=x reference line to Plot 3 (the "ideal fit" diagonal).
        all_vals = np.concatenate([diag['dX_true'][n] for n in var_names])
        vmin, vmax = float(all_vals.min()), float(all_vals.max())
        p_scatter.line(
            [vmin, vmax], [vmin, vmax],
            color="#e74c3c", line_width=1.5, line_dash="dashed",
            legend_label="ideal",
        )

        for fig in [p_resid, p_fft, p_scatter]:
            fig.legend.click_policy = "hide"
            fig.legend.location = "top_right"

        # Build the stats summary — one line per state variable.
        stats_html = "<b>Residual Stats:</b><br>"
        for name, s in diag['stats'].items():
            stats_html += (
                f"&nbsp;&nbsp;<b>{name}</b>: "
                f"R²(dX)={s['r2_dx']} | "
                f"SNR={s['snr_db']} dB | "
                f"autocorr={s['autocorr']}<br>"
            )
        diag_stats_div.text = stats_html

    # =========================================================================
    # SECTION 7 — TRAIN CALLBACK
    # Main entry point triggered by the "TRAIN" button. Loads data (pre-set
    # or uploaded), fits SINDy on a random split, computes diagnostics,
    # updates the leaderboard, and re-renders all plots.
    # =========================================================================

    def on_train_click():
        # ── 1. Resolve the data source (pre-set file vs uploaded CSV) ──────
        is_custom = (file_select.value == "custom_upload")
        uploaded_value = None
        if is_custom:
            try:
                uploaded_value = file_input.value
            except Exception:
                uploaded_value = None

        if is_custom:
            if not uploaded_value:
                # Guard against pressing Train before a file was actually chosen.
                res_div.text = "<span style='color:red;'>⚠ Please upload a CSV file first!</span>"
                return
            # Decode the uploaded CSV (base64 -> bytes -> DataFrame).
            decoded = base64.b64decode(file_input.value)
            f = io.BytesIO(decoded)
            df = pd.read_csv(f).astype(np.float64)
            # FileInput.filename is only populated in newer Bokeh versions —
            # fall back to a generic label so the leaderboard never shows blank.
            data_file_label = getattr(
                file_input, 'filename', None) or "Custom Upload"
        else:
            # Load one of the bundled pre-set system files.
            path = os.path.join('data', file_select.value)
            df = pd.read_csv(path).astype(np.float64)
            data_file_label = _SYSTEM_LABELS.get(
                file_select.value, file_select.value)

        counter[0] += 1  # unique, ever-increasing run ID

        # ── 2. Parse data into time / state matrices ────────────────────────
        t = df.iloc[:, 0].values
        X = df.iloc[:, 1:].values
        names = list(df.columns[1:])
        train_frac = train_s.value / 100.0

        # ── 3. Fit SINDy on a random train/validation split ─────────────────
        # Derivatives are computed once on the FULL trajectory, then the
        # resulting (X, dX) pairs are split randomly — this is more robust
        # than splitting the raw time series first, because finite-difference
        # derivatives near a split boundary would otherwise be biased.
        try:
            model, train_idx, val_idx, m_train, m_val = \
                engine.fit_model(
                    X, t,
                    poly_degree=poly_s.value,
                    threshold=thr_s.value,
                    names=names,
                    lib_type=library_select.value,
                    train_frac=train_frac,
                    random_seed=counter[0] * 7,  # unique seed per run
                    split_method=split_select.value.lower()
                )
        except Exception as e:
            res_div.text = f"<span style='color:red;'>⚠ Fit error: {e}</span>"
            return

        # ── 4. Compute residual diagnostics for the 3 diagnostic plots ──────
        diag = engine.compute_diagnostics(X, t)

        t_r2, t_rmse, t_mae = m_train['r2'], m_train['rmse'], m_train['mae']
        v_r2, v_rmse, v_mae = m_val['r2'],   m_val['rmse'],   m_val['mae']
        rmse_diff = float(np.abs(t_rmse - v_rmse))

        # ── 5. Forward-simulate the discovered equations over the full
        #        time range for visualization (x(t) reconstruction from
        #        the initial condition, NOT the raw dX/dt fit) ─────────────
        try:
            x_sim_full = engine.simulate(X[0], t)
        except Exception as e:
            res_div.text = f"<span style='color:red;'>⚠ Simulation error: {e}</span>"
            return

        # ── 6. Format the discovered equations for display ─────────────────
        raw_eqs = engine.get_equations()
        formatted_eqs_html = "".join(
            [f"<b style='color:#e74c3c;'>({i+1})</b> {eq}<br>" for i,
             eq in enumerate(raw_eqs)]
        )

        # ── 7. Append a new row to the leaderboard ──────────────────────────
        new_entry = {
            'run':        [counter[0]],
            'system':     [data_file_label],
            'split':      [split_select.value],
            'lib':        [library_select.value],
            'poly':       [poly_s.value],
            'thr':        [thr_s.value],
            'train_metrics': [_fmt_metrics_html("TRAIN", "#1f77b4", t_r2, t_rmse, t_mae)],
            'val_metrics':   [_fmt_metrics_html("VAL", "#ff7f0e", v_r2, v_rmse, v_mae)],
            'rmse_diff':  [f"{rmse_diff:.6f}"],
            'equations':  [formatted_eqs_html],
        }
        source_history.stream(new_entry)

        # ── 7b. Mirror this run into the leaderboard log (Excel or Google
        #         Sheet, per LOG_BACKEND) — same columns/order as above, so
        #         results survive app restarts and can be opened outside Bokeh.
        export_run_to_log(
            run_id=counter[0],
            data_file=data_file_label,
            split_type=split_select.value,
            lib=library_select.value,
            poly_degree=poly_s.value,
            threshold=thr_s.value,
            train_r2=t_r2, train_rmse=t_rmse, train_mae=t_mae,
            val_r2=v_r2, val_rmse=v_rmse, val_mae=v_mae,
            rmse_diff=rmse_diff,
            equations_html=formatted_eqs_html,
        )

        # ── 8. Persist everything needed to reconstruct this run later ─────
        # (used by render_plot/_render_diag_plots on row-select, and by the
        # Test/Predict tabs to simulate from a saved model instance).
        trained_model_storage[counter[0]] = {
            'run_id':             counter[0],
            'system_name':        file_select.value,
            'split_strategy':     split_select.value,
            # snapshot — engine.model gets overwritten on next Train
            'model_instance':     copy.deepcopy(engine.model),
            'lib_type':           library_select.value,
            'poly_degree':        poly_s.value,
            'threshold':          thr_s.value,
            'feature_names':      names,             # variable names from the CSV header
            'initial_conditions': X[0].tolist(),
            'metrics': {
                'train_rmse': t_rmse,
                'val_rmse':   v_rmse,
                'rmse_diff':  rmse_diff,
                'val_r2':     v_r2,
            },
            'equations':  raw_eqs,
            'plot_data': {
                't':         t,
                'X':         X,
                'train_idx': train_idx,
                'val_idx':   val_idx,
                'x_sim':     x_sim_full,
            },
            'diagnostics': diag,
        }

        # ── 9. Render the plots for the run that was just trained ──────────
        render_plot(counter[0])
        _render_diag_plots(diag)

    def on_delete_click():
        """
        Remove the currently-selected leaderboard row: deletes the model
        from trained_model_storage, removes the row from the DataTable,
        and clears the main/diagnostic plots if the deleted run was the
        one currently being viewed.

        NOTE: Run IDs (`counter`) are intentionally NOT reset/renumbered
        after a deletion — every run ID stays permanently unique so past
        references (e.g. from the Test/Predict tabs) never become ambiguous.
        """
        selected = source_history.selected.indices
        if not selected:
            return

        idx = selected[0]
        run_id = source_history.data['run'][idx]

        # Remove from the model storage dict.
        if run_id in trained_model_storage:
            del trained_model_storage[run_id]

        # Remove the row from the DataTable by rebuilding every column list
        # with that index filtered out (ColumnDataSource has no native
        # "delete row" API).
        new_data = {k: [v for i, v in enumerate(vals) if i != idx]
                    for k, vals in source_history.data.items()}
        source_history.data = new_data
        source_history.selected.indices = []

        # If the deleted run was the one currently displayed, clear the
        # main result plot back to an empty state.
        if view_div.text and f"Run #{run_id}" in view_div.text:
            p.renderers = []
            if p.legend:
                p.legend.items = []
            p.title.text = "Model Result"
            view_div.text = ""

        # Clear all 3 diagnostic plots too.
        for figs in [p_resid, p_fft, p_scatter]:
            figs.renderers = []
            if figs.legend and len(figs.legend) > 0:
                figs.legend[0].items = []

        # Reset stats text and FFT x-axis range back to neutral defaults —
        # the range will be auto-scaled again on the next training run.
        diag_stats_div.text = "<i>Run a training session to see diagnostics.</i>"
        p_fft.x_range.start = 0.0
        p_fft.x_range.end = 1.0

    btn_delete.on_click(on_delete_click)

    # ── Run the AI Suggester once on page load for the default pre-set
    #     system, so the sliders aren't left at arbitrary defaults before
    #     the user has interacted with anything. ──────────────────────────
    initial_path = os.path.join('data', file_select.value)
    if os.path.exists(initial_path):
        try:
            df_init = pd.read_csv(initial_path).astype(np.float64)
            apply_suggestion(
                df_init, f"<b>Loaded default pre-set system: {file_select.value}</b>")
        except Exception:
            pass  # non-fatal — user can still configure manually

    btn_train.on_click(on_train_click)

    # =========================================================================
    # SECTION 8 — LAYOUT ASSEMBLY
    # =========================================================================

    top_row = row(
        column(file_select, file_input, train_s, split_select, library_select,
               poly_s, thr_s, thr_input, row(btn_train, btn_delete), width=320),
        column(p, view_div, sizing_mode="stretch_width"),
        sizing_mode="stretch_width"
    )

    return column(
        top_row,
        Div(text="<hr><b>RESIDUAL DIAGNOSTICS</b>"),
        diag_stats_div,
        row(p_resid, p_fft, p_scatter, sizing_mode="stretch_width"),
        Div(text="<hr><b>TRAINING HISTORY — Metrics on dx/dt (derivative space)</b>"),
        history_table,
        sizing_mode="stretch_width"
    )